# maintenance_alert_refactored.py
import pandas as pd
from datetime import datetime, timedelta, date
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
import sys
import json
import shutil
import matplotlib
matplotlib.use('Agg')  # Use non-GUI backend for thread safety
import matplotlib.pyplot as plt
import logging
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional, Any, NamedTuple

# --- 1. Конфигурация и константы ---
class Config:
    """Класс для хранения всех конфигурационных данных."""
    VERSION = "2.7.1"
    RELEASE_DATE = "02.05.2026"

    PROGRAM_DIR = Path(__file__).parent.absolute()
    DATA_DIR = PROGRAM_DIR / "data"
    TMP_DIR = PROGRAM_DIR / "tmp"
    BACKUP_DIR = PROGRAM_DIR / "backups_excel"
    LOG_FILE = DATA_DIR / "maintenance_alert.log"

    EXCEL_FILENAME = "Обслуживание ПК и шкафов АСУТП.xlsx"
    HISTORY_FILE = DATA_DIR / "maintenance_alert_history.json"
    SERVICED_HISTORY_FILE = DATA_DIR / "serviced_history.json"
    EXCEL_SNAPSHOT_FILE = DATA_DIR / "excel_snapshot.json"

    SMTP_SERVER = "mgd-ex1.pavlik-gold.ru"
    SMTP_PORT = 25
    SENDER_EMAIL = "maintenance.asutp@pavlik-gold.ru"
    RECIPIENTS = [
        "asutp@pavlik-gold.ru",
        # "ochkur.evgeniy@pavlik-gold.ru",
        # "dorovik.roman@pavlik-gold.ru",
    ]

    COLUMN_NAMES = [
        "№", "Объект", "Наименование", "Обозначение", "Место расположения",
        "Работы", "Интервал ТО (дней)", "Напоминание (за дней)", "Дата последнего ТО",
        "Дата следующего ТО", "Статус"
    ]
    SHEETS_CONFIG = {
        "ПК АСУ ТП": {"range": "A4:K300"},
        "Шкафы АСУ ТП": {"range": "A4:K300"}
    }
    MAINTENANCE_STATUSES = ["ОБСЛУЖИТЬ", "Внимание", "Не требуется"]
    STATUS_URGENT = "ОБСЛУЖИТЬ"
    STATUS_WARNING = "Внимание"
    STATUS_OK = "Не требуется"
    
    CHART_DAYS = 62  # Количество дней отображаемых в диаграмме
    HISTORY_MAX_DAYS = 180  # Глубина хранения истории обслуживания (6 месяцев ≈ 180 дней)

    @classmethod
    def get_excel_file_path(cls) -> Path:
        """Ищет Excel-файл сначала в папке скрипта, затем уровнем выше."""
        primary = cls.PROGRAM_DIR / cls.EXCEL_FILENAME
        if primary.exists():
            return primary
        fallback = cls.PROGRAM_DIR.parent / cls.EXCEL_FILENAME
        if fallback.exists():
            return fallback
        return primary # Возвращаем путь в папке скрипта для ошибки

# --- 2. Логирование ---
class DualLogger:
    """Класс для дублированного вывода в консоль и файл."""
    def __init__(self, log_file_path: Path):
        self.log_file_path = log_file_path
        self._setup_logging()

    def _setup_logging(self):
        """Настройка системы логирования."""
        self.log_file_path.parent.mkdir(parents=True, exist_ok=True)
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S',
            handlers=[
                logging.FileHandler(self.log_file_path, encoding='utf-8'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        self.logger = logging.getLogger('maintenance_alert')

    def log(self, message: str):
        """Выводит сообщение в консоль и записывает в файл."""
        self.logger.info(message)

    def log_separator(self, char='=', length=60):
        """Добавляет разделитель в лог."""
        self.log(char * length)

    def log_section(self, title: str):
        """Добавляет заголовок секции в лог."""
        self.log_separator()
        self.log(title)
        self.log_separator()

# --- 3. Работа с файлами и Excel ---
class ExcelHandler:
    """Класс для работы с Excel файлами."""
    def __init__(self, config: Config, logger: DualLogger):
        self.config = config
        self.logger = logger
        self.xlwings_available = self._check_xlwings()

    def _check_xlwings(self) -> bool:
        """Проверяет доступность xlwings."""
        try:
            import xlwings as xw
            self.xw = xw
            return True
        except ImportError:
            self.logger.log("⚠️ xlwings недоступен. Формулы Excel могут быть неактуальными.")
            self.logger.log("💡 Установите: pip install xlwings")
            return False

    def _verify_file_write(self, file_path: Path, original_mtime: float = None) -> bool:
        """Проверяет, что файл был успешно сохранен и обновлен."""
        try:
            if not file_path.exists() or file_path.stat().st_size == 0:
                return False
            with open(file_path, 'rb') as f:
                if len(f.read(8)) < 8: return False
            if original_mtime is not None:
                current_mtime = file_path.stat().st_mtime
                if current_mtime <= original_mtime:
                    self.logger.log(f"⚠️ Файл не был обновлен: ориг. {original_mtime:.1f}, тек. {current_mtime:.1f}")
                    return False
                else:
                    self.logger.log(f"✅ Файл обновлен: разница {current_mtime - original_mtime:.1f} сек")
            return True
        except Exception as e:
            self.logger.log(f"❌ Ошибка проверки файла: {e}")
            return False

    def recalculate_formulas(self, file_path: Path) -> Tuple[bool, Optional[Path]]:
        """Пересчитывает формулы в Excel файле."""
        if not self.xlwings_available:
            return False, None
        if not file_path.exists():
            self.logger.log(f"❌ Файл не найден: {file_path}")
            return False, None

        self.config.TMP_DIR.mkdir(parents=True, exist_ok=True)
        tmp_file_path = self.config.TMP_DIR / file_path.name

        try:
            self.logger.log(f"🔄 Пересчитываем формулы с xlwings: {file_path}")
            original_mtime = file_path.stat().st_mtime

            with self.xw.App(visible=False, add_book=False) as app:
                wb = app.books.open(str(file_path))
                try:
                    app.calculation = 'automatic'
                    wb.app.calculate()
                    for sheet in wb.sheets:
                        if sheet.name in self.config.SHEETS_CONFIG:
                            try:
                                sheet.api.Calculate()
                            except AttributeError:
                                pass
                    wb.save(tmp_file_path)

                    if not self._verify_file_write(tmp_file_path):
                        self.logger.log("❌ Ошибка: файл не был корректно сохранен в tmp папку!")
                        return False, None

                    self.logger.log(f"✅ Формулы успешно пересчитаны и сохранены в {tmp_file_path}")
                    return True, tmp_file_path
                finally:
                    wb.close()
        except Exception as e:
            self.logger.log(f"❌ Ошибка при пересчете с xlwings: {e}")
            self.logger.log("💡 Совет: убедитесь, что файл Excel не открыт в другом приложении")
            return False, None

    def generate_maintenance_data_file(self, urgent_items: List[pd.DataFrame]) -> Optional[Path]:
        """
        Создает файл maintenance_data.xlsx на основе шаблона с данными для обслуживания.
        Args:
            urgent_items: Список DataFrame с элементами требующими обслуживания
        Returns:
            Путь к созданному файлу или None при ошибке
        """
        # Получаем текущую дату в формате DD.MM.YYYY
        current_date = datetime.now().strftime("%d.%m.%Y")
        # current_date2 = datetime.now().strftime("%d_%m_%Y")
        template_path = self.config.DATA_DIR / "template.xlsx"
        output_path = self.config.TMP_DIR / f"maintenance_data_{current_date}.xlsx"
        
        if not template_path.exists():
            self.logger.log(f"❌ Шаблон не найден: {template_path}")
            return None
            
        try:
            # Создаем папку tmp если не существует
            self.config.TMP_DIR.mkdir(parents=True, exist_ok=True)
            
            # Копируем шаблон
            wb = load_workbook(template_path)
            
            
            # Определяем столбцы для экспорта
            export_columns = ["№", "Объект", "Наименование", "Обозначение", "Место расположения", 
            "Работы", "Интервал ТО (дней)", "Напоминание (за дней)", "Дата последнего ТО", "Дата следующего ТО", "Статус" ]
            
            # Обрабатываем каждый лист
            for sheet_name in self.config.SHEETS_CONFIG.keys():
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Записываем дату в ячейку D1
                    ws['D1'] = current_date
                    self.logger.log(f"📅 Записана дата {current_date} в ячейку D1 листа '{sheet_name}'")
                    
                    # Находим данные для этого листа
                    sheet_data = None
                    for df in urgent_items:
                        if 'Тип' in df.columns and df['Тип'].iloc[0] == sheet_name:
                            sheet_data = df
                            break
                    
                    if sheet_data is not None and not sheet_data.empty:
                        self.logger.log(f"📝 Записываем {len(sheet_data)} записей на лист '{sheet_name}'")
                        ws['D2'] = len(sheet_data)

                        # Записываем данные начиная с 5й строки
                        start_row = 5
                        record_number = 1  # Нумерация записей начинается с 1
                        for idx, (_, row) in enumerate(sheet_data.iterrows()):
                            current_row = start_row + idx
                            
                            # Записываем данные в соответствующие столбцы
                            for col_idx, col_name in enumerate(export_columns, start=1):
                                if col_name == "№":
                                    # Для колонки "№" используем последовательную нумерацию
                                    value = str(record_number)
                                elif col_name in row:
                                    value = row[col_name]
                                    # Преобразуем в скаляр
                                    if hasattr(value, 'item'):
                                        value = value.item()
                                    # Форматируем дату если это колонка с датой
                                    if 'Дата' in col_name and pd.notna(value):
                                        if hasattr(value, 'strftime'):
                                            value = value.strftime('%d.%m.%Y')
                                        else:
                                            value = str(value)
                                    elif pd.isna(value):
                                        value = ""
                                    else:
                                        value = str(value) if value is not None else ""
                                else:
                                    value = ""
                                
                                ws.cell(row=current_row, column=col_idx, value=value)
                            
                            record_number += 1  # Увеличиваем номер записи
                    else:
                        self.logger.log(f"ℹ️ Нет данных для записи на лист '{sheet_name}'")

                    #Устанавливаем фокус  в левый верхний угол
                    # wb.views.sheetView[0].topLeftCell = 'A1'
                    # ws['A1'].select()


            # Сохраняем файл
            wb.save(output_path)
            wb.close()
            
            self.logger.log(f"✅ Файл maintenance_data.xlsx создан: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.log(f"❌ Ошибка при создании файла maintenance_data.xlsx: {e}")
            return None

    def read_data(self) -> Tuple[List[pd.DataFrame], List[pd.DataFrame], int, Dict[str, int], bool]:
        """Читает данные из Excel файла."""
        recalc_success, excel_file_to_use = self.recalculate_formulas(self.config.get_excel_file_path())

        if excel_file_to_use is None:
            excel_file_to_use = self.config.get_excel_file_path()
            self.logger.log(f"⚠️ Используем оригинальный файл: {excel_file_to_use}")
        else:
            self.logger.log(f"✅ Используем файл с пересчитанными формулами: {excel_file_to_use}")

        urgent_items = []
        warning_items = []
        total_records = 0
        status_counts = {status: 0 for status in self.config.MAINTENANCE_STATUSES}

        for sheet_name, config in self.config.SHEETS_CONFIG.items():
            try:
                self.logger.log(f"Читаем лист: {sheet_name}")
                df = pd.read_excel(excel_file_to_use, sheet_name=sheet_name, header=3, nrows=500)

                if len(df.columns) > len(self.config.COLUMN_NAMES):
                    df = df.iloc[:, :len(self.config.COLUMN_NAMES)]
                df.columns = self.config.COLUMN_NAMES
                df = df.dropna(how='all')

                total_records += len(df)
                for status in status_counts.keys():
                    status_counts[status] += len(df[df['Статус'] == status])

                urgent_df = df[df['Статус'] == self.config.STATUS_URGENT]
                warning_df = df[df['Статус'] == self.config.STATUS_WARNING]

                self.logger.log(f"  Найдено {self.config.STATUS_URGENT}: {len(urgent_df)}, {self.config.STATUS_WARNING}: {len(warning_df)}")

                if not urgent_df.empty:
                    urgent_df = urgent_df.copy()
                    urgent_df['Тип'] = sheet_name
                    urgent_items.append(urgent_df)
                if not warning_df.empty:
                    warning_df = warning_df.copy()
                    warning_df['Тип'] = sheet_name
                    warning_items.append(warning_df)
            except Exception as e:
                self.logger.log(f"Ошибка при чтении листа {sheet_name}: {e}")

        # Сохраняем путь к файлу для последующего использования
        self.last_excel_file_path = excel_file_to_use
        return urgent_items, warning_items, total_records, status_counts, recalc_success

    def get_last_excel_file_path(self) -> Path:
        """Возвращает путь к последнему использованному Excel файлу."""
        return getattr(self, 'last_excel_file_path', self.config.get_excel_file_path())

    def is_file_locked(self, file_path: Path) -> bool:
        """
        Проверяет, заблокирован ли файл другим приложением (открыт в Excel).
        На Windows это делается попыткой переименовать файл.
        """
        if not file_path.exists():
            return False
        try:
            # Попытка открыть файл в эксклюзивном режиме для записи
            # Если файл открыт в Excel, это вызовет PermissionError
            f = open(file_path, 'a')
            f.close()
            return False
        except (IOError, PermissionError):
            return True

    def create_backup(self, file_path: Path) -> Tuple[bool, str]:
        """
        Создает резервную копию файла в папке backups_excel.
        Добавляет к имени файла текущую дату и время.
        """
        try:
            if not file_path.exists():
                return False, "Исходный файл не найден"
            
            self.config.BACKUP_DIR.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            backup_filename = f"{file_path.stem}_{timestamp}{file_path.suffix}"
            backup_path = self.config.BACKUP_DIR / backup_filename
            
            shutil.copy2(file_path, backup_path)
            self.logger.log(f"💾 Резервная копия создана: {backup_path.name}")
            return True, str(backup_path)
        except Exception as e:
            msg = f"Ошибка при создании резервной копии: {e}"
            self.logger.log(f"⚠️ {msg}")
            return False, msg

    def mark_as_serviced(self, sheet_name: str, row_number: str, make_backup: bool = True) -> Tuple[bool, str]:
        """
        Отмечает оборудование как обслуженное, обновляя дату последнего ТО.
        
        Args:
            sheet_name: Название листа Excel ("ПК АСУ ТП" или "Шкафы АСУ ТП")
            row_number: Номер строки из колонки № (колонка A) для поиска записи
            make_backup: Нужно ли создавать резервную копию перед изменением
            
        Returns:
            Tuple[bool, str]: (успех операции, сообщение)
        """
        file_path = self.config.get_excel_file_path()
        if not file_path.exists():
            return False, f"Файл не найден: {file_path}"
            
        # Проверка на блокировку файла перед началом работы
        if self.is_file_locked(file_path):
            return False, "⚠️ Файл Excel открыт в другой программе! Закройте его перед выполнением операции."
            
        # Создание резервной копии перед изменением
        if make_backup:
            self.create_backup(file_path)
            
        try:
            self.logger.log(f"📝 Отмечаем как обслуженное: {sheet_name} / №{row_number}")
            
            # Открываем Excel файл с помощью openpyxl
            wb = load_workbook(str(file_path))
            
            try:
                # Проверяем наличие листа
                if sheet_name not in wb.sheetnames:
                    return False, f"Лист '{sheet_name}' не найден в файле"
                
                sheet = wb[sheet_name]
                
                # Ищем строку с нужным номером в колонке A (№)
                # Колонка A (№) - индекс 1
                # Колонка I (Дата последнего ТО) - индекс 9
                start_row = 5  # Данные начинаются со строки 5 (строка 4 - заголовки)
                max_rows = 500
                found_row = None
                
                # Преобразуем row_number в число для сравнения
                try:
                    target_number = int(row_number)
                except ValueError:
                    return False, f"Некорректный номер строки: '{row_number}'"
                
                for row_idx in range(start_row, start_row + max_rows):
                    cell_value = sheet.cell(row=row_idx, column=1).value  # Колонка A (№)
                    if cell_value is not None:
                        try:
                            # Преобразуем значение ячейки в число
                            cell_number = int(float(cell_value))
                            if cell_number == target_number:
                                found_row = row_idx
                                break
                        except (ValueError, TypeError):
                            continue
                
                if found_row is None:
                    return False, f"Оборудование с номером '№{row_number}' не найдено на листе '{sheet_name}'"
                
                # Получаем обозначение для лога (колонка D = 4)
                designation = sheet.cell(row=found_row, column=4).value or "N/A"
                
                # Обновляем дату последнего ТО (колонка I = 9)
                today = datetime.now()
                sheet.cell(row=found_row, column=9).value = today
                
                # Сохраняем файл
                wb.save(str(file_path))
                
                self.logger.log(f"✅ Успешно обновлена дата ТО для '№{row_number} ({designation})' в строке {found_row}")
                return True, f"Оборудование '№{row_number} ({designation})' отмечено как обслуженное"
                
            finally:
                wb.close()
                
        except Exception as e:
            error_msg = f"Ошибка при обновлении: {str(e)}"
            self.logger.log(f"❌ {error_msg}")
            return False, error_msg

# --- 4. Логика обслуживания ---
class MaintenanceChecker:
    """Класс для анализа статусов обслуживания."""
    def __init__(self, config: Config, logger: DualLogger):
        self.config = config
        self.logger = logger

    def format_date(self, date_value) -> str:
        """Форматирует дату в формат dd.mm.yyyy."""
        if pd.notna(date_value):
            # Проверяем на тип datetime или date
            if hasattr(date_value, 'strftime'):
                try:
                    return date_value.strftime('%d.%m.%Y')
                except (ValueError, AttributeError):
                    pass # Если strftime не удался, продолжаем
            # Если это строка или число, пытаемся преобразовать
            return str(date_value)
        return "Не указана"

    def format_field_value(self, value) -> str:
        """Форматирует значение поля, обрабатывая NaN значения."""
        if pd.isna(value):
            return ""
        return str(value)

    def format_item_info(self, item: pd.Series, item_type: str) -> str:
        """Форматирует информацию об элементе."""
        emoji = "💻" if "ПК" in item_type else ("📦" if "Шкаф" in item_type else "⚙️")

        raboty_row = ""
        if not pd.isna(item['Работы']):
            raboty_value = self.format_field_value(item['Работы'])
            raboty_row = f"<tr><td style='padding: 1px 10px 1px 0; width: 200px; color:#2c3e50; vertical-align: top;'>Работы:</td><td style='padding: 1px 0; color:#2c3e50; font-weight: bold;'>{raboty_value}</td></tr>"

        return f"""
<div style='margin-bottom: 10px;'>
    <table style='width: 100%; border-collapse: collapse; font-size: 14px;'>
        <tr><td style='padding: 1px 10px 1px 0; width: 200px; color:#2c3e50; vertical-align: top;'>Тип:</td><td style='padding: 1px 0; color:#2c3e50; font-weight: bold;'>{emoji}  {item_type}</td></tr>
        <tr><td style='padding: 1px 10px 1px 0; width: 200px; color:#2c3e50; vertical-align: top;'>Объект:</td><td style='padding: 1px 0; color:#2c3e50; font-weight: bold;'>{item['Объект']}</td></tr>
        <tr><td style='padding: 1px 10px 1px 0; width: 200px; color:#2c3e50; vertical-align: top;'>Наименование:</td><td style='padding: 1px 0; color:#2c3e50; font-weight: bold;'>{item['Наименование']}</td></tr>
        <tr><td style='padding: 1px 10px 1px 0; width: 200px; color:#2c3e50; vertical-align: top;'>Обозначение:</td><td style='padding: 1px 0; color:#2c3e50; font-weight: bold;'>{item['Обозначение']}</td></tr>
        <tr><td style='padding: 1px 10px 1px 0; width: 200px; color:#2c3e50; vertical-align: top;'>Место расположения:</td><td style='padding: 1px 0; color:#2c3e50; font-weight: bold;'>{item['Место расположения']}</td></tr>
        {raboty_row}
        <tr><td style='padding: 1px 10px 1px 0; width: 200px; color:#2c3e50; vertical-align: top;'>Интервал ТО (дней):</td><td style='padding: 1px 0; color:#2c3e50; font-weight: bold;'>{item['Интервал ТО (дней)']}</td></tr>
        <tr><td style='padding: 1px 10px 1px 0; width: 200px; color:#2c3e50; vertical-align: top;'>Дата последнего ТО:</td><td style='padding: 1px 0; color:#2c3e50; font-weight: bold;'>{self.format_date(item['Дата последнего ТО'])}</td></tr>
        <tr><td style='padding: 1px 10px 1px 0; width: 200px; color:#2c3e50; vertical-align: top;'>Дата следующего ТО:</td><td style='padding: 1px 0; color:#2c3e50; font-weight: bold;'>{self.format_date(item['Дата следующего ТО'])}</td></tr>
        <tr><td style='padding: 1px 10px 1px 0; width: 200px; color:#2c3e50; vertical-align: top;'>Статус:</td><td style='padding: 1px 0; color:#2c3e50; font-weight: bold;'>{item['Статус']}</td></tr>
    </table>
</div>
"""

    def format_item_table_row(self, item: pd.Series, bg_color: str) -> str:
        """Форматирует строку таблицы для элемента (аналогично web-интерфейсу)."""
        # Получаем значения, обрабатывая NaN
        works = self.format_field_value(item['Работы']) if not pd.isna(item['Работы']) else ''
        
        # Форматируем интервал как целое число
        interval_days = int(item['Интервал ТО (дней)']) if not pd.isna(item['Интервал ТО (дней)']) else ''
        
        # Определяем цвет статуса
        status = item['Статус']
        status_color = '#e74c3c' if status == self.config.STATUS_URGENT else '#f39c12'
        
        return f"""
                    <tr style='background-color: {bg_color};'>
                        <td style='padding:8px; border:1px solid #cfd8dc;'>{item['Объект']}</td>
                        <td style='padding:8px; border:1px solid #cfd8dc;'>{item['Наименование']}</td>
                        <td style='padding:8px; border:1px solid #cfd8dc;'><strong>{item['Обозначение']}</strong></td>
                        <td style='padding:8px; border:1px solid #cfd8dc;'>{item['Место расположения']}</td>
                        <td style='padding:8px; border:1px solid #cfd8dc;'>{works}</td>
                        <td style='padding:8px; border:1px solid #cfd8dc;'>{interval_days}</td>
                        <td style='padding:8px; border:1px solid #cfd8dc;'>{self.format_date(item['Дата последнего ТО'])}</td>
                        <td style='padding:8px; border:1px solid #cfd8dc;'>{self.format_date(item['Дата следующего ТО'])}</td>
                        <td style='padding:8px; border:1px solid #cfd8dc;'><div style='font-weight:bold; color:{status_color};'>{status}</div></td>
                    </tr>
"""

# --- 5. Статистика ---
class StatisticsManager:
    """Класс для управления статистикой обслуживания."""
    def __init__(self, config: Config, logger: DualLogger):
        self.config = config
        self.logger = logger
        self.history_file = self.config.HISTORY_FILE

    def load_config(self) -> Dict[str, Any]:
        """Загружает конфигурацию из JSON файла."""
        try:
            if self.history_file.exists():
                with open(self.history_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    return self._validate_config_structure(config)
            else:
                return self._create_default_config()
        except Exception as e:
            self.logger.log(f"Ошибка при загрузке конфигурации: {e}")
            return self._create_default_config()

    def _validate_config_structure(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """Проверяет и корректирует структуру конфигурации."""
        if 'maintenance_history' not in config:
            config['maintenance_history'] = []
        if 'last_update' not in config:
            config['last_update'] = None
        if 'version' not in config:
            config['version'] = self.config.VERSION
        return config

    def _create_default_config(self) -> Dict[str, Any]:
        """Создает конфигурацию по умолчанию."""
        config = {
            "maintenance_history": [],
            "last_update": None,
            "version": self.config.VERSION
        }
        self.save_config(config)
        return config

    def save_config(self, config: Dict[str, Any]) -> None:
        """Сохраняет конфигурацию в JSON файл."""
        try:
            config['last_update'] = datetime.now().isoformat()
            config['version'] = self.config.VERSION
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            self.logger.log(f"✅ Статистика сохранена в {self.history_file}")
        except Exception as e:
            self.logger.log(f"❌ Ошибка при сохранении конфигурации: {e}")

    def update_statistics(self, urgent_items: List[pd.DataFrame],
                          warning_items: List[pd.DataFrame],
                          total_records: int,
                          status_counts: Dict[str, int]) -> Dict[str, Any]:
        """Обновляет статистику обслуживания."""
        config = self.load_config()
        today = datetime.now().date()
        today_str = today.isoformat()
        self.logger.log(f"🔍 Проверяем существование записи за {today.strftime('%d.%m.%Y')}...")

        ok_count = status_counts.get(self.config.STATUS_OK, 0)
        maintenance_record = {
            "date": today_str,
            "total_equipment": total_records,
            "ok": ok_count,
            "urgent": status_counts.get(self.config.STATUS_URGENT, 0),
            "warning": status_counts.get(self.config.STATUS_WARNING, 0),
            "timestamp": datetime.now().isoformat()
        }

        today_record_index = next((i for i, record in enumerate(config['maintenance_history']) if record['date'] == today_str), None)

        try:
            if today_record_index is not None:
                self.logger.log(f"📝 Перезаписываем существующую запись за {today.strftime('%d.%m.%Y')}...")
                config['maintenance_history'][today_record_index] = maintenance_record
                action = "обновлена"
            else:
                self.logger.log(f"📝 Создаем новую запись за {today.strftime('%d.%m.%Y')}...")
                config['maintenance_history'].append(maintenance_record)
                action = "добавлена"

            if len(config['maintenance_history']) > self.config.HISTORY_MAX_DAYS + 5:
                config['maintenance_history'] = config['maintenance_history'][-self.config.HISTORY_MAX_DAYS:]

            self.save_config(config)
            self.logger.log(f"✅ Запись за {today.strftime('%d.%m.%Y')} {action}: {ok_count} обслужено")
            return config
        except Exception as e:
            self.logger.log(f"❌ Ошибка при обновлении статистики: {e}")
            return config

    def _compute_period_boundaries(self, base_date: date) -> Dict[str, date]:
        """Вычисляет границы периодов для статистики."""
        yesterday_local = base_date - timedelta(days=1)
        day_before_yesterday_local = yesterday_local - timedelta(days=1)
        week_start_local = base_date - timedelta(days=base_date.weekday())
        last_week_start_local = week_start_local - timedelta(days=7)
        last_week_end_local = week_start_local - timedelta(days=1)
        prev_prev_week_start_local = last_week_start_local - timedelta(days=7)
        prev_prev_week_end_local = last_week_start_local - timedelta(days=1)
        month_start_local = base_date.replace(day=1)
        last_month_end_local = month_start_local - timedelta(days=1)
        last_month_start_local = last_month_end_local.replace(day=1)
        prev_prev_month_end_local = last_month_start_local - timedelta(days=1)
        prev_prev_month_start_local = prev_prev_month_end_local.replace(day=1)
        return {
            "yesterday": yesterday_local,
            "day_before_yesterday": day_before_yesterday_local,
            "week_start": week_start_local,
            "last_week_start": last_week_start_local,
            "last_week_end": last_week_end_local,
            "prev_prev_week_start": prev_prev_week_start_local,
            "prev_prev_week_end": prev_prev_week_end_local,
            "month_start": month_start_local,
            "last_month_start": last_month_start_local,
            "last_month_end": last_month_end_local,
            "prev_prev_month_start": prev_prev_month_start_local,
            "prev_prev_month_end": prev_prev_month_end_local,
        }

    def _aggregate_raw_field(self, history_records: List[Dict],
                            today_local: date,
                            bounds: Dict[str, date],
                            extract_value) -> Dict[str, int]:
        """Агрегирует данные по периодам."""
        raw = {
            "today": 0,
            "yesterday": 0,
            "day_before_yesterday": 0,
            "this_week": 0,
            "last_week": 0,
            "week_before_last": 0,
            "this_month": 0,
            "last_month": 0,
            "month_before_last": 0,
        }
        for record in history_records:
            record_date = datetime.fromisoformat(record['date']).date()
            value = extract_value(record)
            if record_date == today_local:
                raw["today"] = value
            elif record_date == bounds["yesterday"]:
                raw["yesterday"] = value
            elif record_date == bounds["day_before_yesterday"]:
                raw["day_before_yesterday"] = value
            elif bounds["week_start"] <= record_date <= today_local:
                raw["this_week"] = max(raw["this_week"], value)
            elif bounds["last_week_start"] <= record_date <= bounds["last_week_end"]:
                raw["last_week"] = max(raw["last_week"], value)
            elif bounds["prev_prev_week_start"] <= record_date <= bounds["prev_prev_week_end"]:
                raw["week_before_last"] = max(raw["week_before_last"], value)
            elif bounds["month_start"] <= record_date <= today_local:
                raw["this_month"] = max(raw["this_month"], value)
            elif bounds["last_month_start"] <= record_date <= bounds["last_month_end"]:
                raw["last_month"] = max(raw["last_month"], value)
            elif bounds["prev_prev_month_start"] <= record_date <= bounds["prev_prev_month_end"]:
                raw["month_before_last"] = max(raw["month_before_last"], value)
        return raw

    def _compute_delta_stats(self, raw_stats: Dict[str, int]) -> Dict[str, int]:
        """Вычисляет дельты для статистики."""
        return {
            "delta_ok_day": raw_stats["today"] - raw_stats["yesterday"],
            "delta_ok_prev_day": raw_stats["yesterday"] - raw_stats["day_before_yesterday"],
            "delta_ok_week": raw_stats["this_week"] - raw_stats["last_week"],
            "delta_ok_prev_week": raw_stats["last_week"] - raw_stats["week_before_last"],
            "delta_ok_month": raw_stats["this_month"] - raw_stats["last_month"],
            "delta_ok_prev_month": raw_stats["last_month"] - raw_stats["month_before_last"],
        }

    def get_statistics(self) -> Dict[str, int]:
        """Получает статистику обслуживания за различные периоды."""
        config = self.load_config()
        if not config['maintenance_history']:
            return {
                "today": 0,
                "yesterday": 0,
                "this_week": 0,
                "last_week": 0,
                "this_month": 0,
                "last_month": 0
            }
        today = datetime.now().date()
        bounds = self._compute_period_boundaries(today)
        # Агрегируем данные для обслуженных и срочных элементов
        ok_raw_stats = self._aggregate_raw_field(
            config['maintenance_history'], today, bounds,
            lambda rec: rec.get('ok', rec.get('serviced', 0))
        )
        urgent_raw_stats = self._aggregate_raw_field(
            config['maintenance_history'], today, bounds,
            lambda rec: rec.get('urgent', 0)
        )
        # Вычисляем дельты
        ok_delta_stats = self._compute_delta_stats(ok_raw_stats)
        urgent_delta_stats = self._compute_delta_stats(urgent_raw_stats)
        # Объединяем все данные
        merged = {
            **ok_raw_stats,
            **ok_delta_stats,
            **{f"urgent_{k}": v for k, v in urgent_raw_stats.items()},
            **urgent_delta_stats
        }
        merged["today"] = merged["delta_ok_day"]
        return merged

    def _add_chart_labels(self, x: List[int],
                         ok_vals: List[int],
                         urgent_vals: List[int],
                         warning_vals: List[int]) -> None:
        """Добавляет подписи значений на диаграмму."""
        for i, xpos in enumerate(x):
            total_val = ok_vals[i] + urgent_vals[i] + warning_vals[i]
            if total_val <= 0:
                continue
            # Подписи для "В норме" (самый нижний слой)
            if ok_vals[i] > 0:
                pct = ok_vals[i] / total_val * 100
                if pct >= 5:
                    y_pos = ok_vals[i] / 2
                    plt.text(
                        xpos, y_pos,
                        f"{ok_vals[i]}",
                        ha='center', va='center', rotation=90, fontsize=6, color='white'
                    )
            # Подписи для "Внимание" (посередине)
            if warning_vals[i] > 0:
                pct = warning_vals[i] / total_val * 100
                if pct >= 5:
                    y_pos = ok_vals[i] + warning_vals[i] / 2
                    plt.text(
                        xpos, y_pos,
                        f"{warning_vals[i]}",
                        ha='center', va='center', rotation=90, fontsize=6, color='black'
                    )
            # Подписи для "СРОЧНО" (сверху)
            if urgent_vals[i] > 0:
                pct = urgent_vals[i] / total_val * 100
                if pct >= 5:
                    y_pos = ok_vals[i] + warning_vals[i] + urgent_vals[i] / 2
                    plt.text(
                        xpos, y_pos,
                        f"{urgent_vals[i]}",
                        ha='center', va='center', rotation=90, fontsize=6, color='white'
                    )

    def create_chart(self, offset_days: int = 0) -> Optional[Path]:
        """Создает диаграмму статусов обслуживания.
        
        Args:
            offset_days: Смещение в днях от текущей даты (отрицательное = назад, положительное = вперед)
        """
        try:
            config = self.load_config()
            if not config['maintenance_history']:
                return None
            today = datetime.now().date() + timedelta(days=offset_days)
            start_date = today - timedelta(days=self.config.CHART_DAYS - 1)
            # Собираем значения за каждый день диапазона
            date_to_vals = {}
            for rec in config['maintenance_history']:
                rec_date = datetime.fromisoformat(rec['date']).date()
                if start_date <= rec_date <= today:
                    date_to_vals[rec_date] = (
                        rec.get('ok', rec.get('serviced', 0)),
                        rec.get('urgent', 0),
                        rec.get('warning', 0),
                    )
            # Подготавливаем данные для графика
            days_sorted = [start_date + timedelta(days=i) for i in range(self.config.CHART_DAYS)]
            ok_vals = [date_to_vals.get(d, (0, 0, 0))[0] for d in days_sorted]
            urgent_vals = [date_to_vals.get(d, (0, 0, 0))[1] for d in days_sorted]
            warning_vals = [date_to_vals.get(d, (0, 0, 0))[2] for d in days_sorted]
            # Создаем график
            x = list(range(len(days_sorted)))
            plt.figure(figsize=(9, 3))
            # Настройка рамки
            ax = plt.gca()
            for spine in ax.spines.values():
                spine.set_color('#2c3e50')
                spine.set_linewidth(0.8)
            # Правильный порядок слоев: снизу вверх
            # 1. "ОБСЛУЖИТЬ" (сверху) - поверх всех
            bottom_stack = [ok_vals[i] + warning_vals[i] for i in range(len(x))]
            urgent_bars = plt.bar(x, urgent_vals, bottom=bottom_stack, width=0.9, color='#e74c3c', label='ОБСЛУЖИТЬ')
            # 2. "Внимание" (посередине) - поверх "В норме"
            warning_bars = plt.bar(x, warning_vals, bottom=ok_vals, width=0.9, color='#f39c12', label='Внимание')
            # 3. "Не требуется" (самый нижний слой)
            ok_bars = plt.bar(x, ok_vals, width=0.9, color='#18bc9c', label='Не требуется')
            # Добавляем подписи значений
            self._add_chart_labels(x, ok_vals, urgent_vals, warning_vals)
            # Настраиваем оси и легенду
            labels = [d.strftime('%d.%m') for d in days_sorted]
            tick_step = max(1, len(x) // 31)
            tick_positions = list(range(0, len(x), tick_step))
            tick_labels = [labels[i] for i in tick_positions]
            plt.xticks(tick_positions, tick_labels, rotation=45, ha='right', fontsize=6, color="#2c3e50")
            plt.yticks(fontsize=6, color="#2c3e50")
            
            # Заголовок с указанием диапазона дат
            title = f'Статусы по дням ({start_date.strftime("%d.%m.%Y")} - {today.strftime("%d.%m.%Y")})'
            plt.title(title, fontsize=7, color="#2c3e50")
            plt.legend(loc='lower left', fontsize=7)
            
            # Устанавливаем границы так, чтобы отступы от краев были чуть больше отступов между столбцами (+2 пикселя)
            # При ширине столбца 0.9, отступ между ними 0.1. 
            # Добавляем еще немного к отступам от краев (примерно 0.1 в единицах данных соответствует ~2 пикселя при текущем DPI)
            plt.xlim(-0.65, len(x) - 0.35)
            
            plt.tight_layout()
            plt.grid(axis='y', linestyle='--', linewidth=0.5, alpha=0.7)
            # Сохраняем диаграмму
            self.config.DATA_DIR.mkdir(parents=True, exist_ok=True)
            chart_path = self.config.DATA_DIR / 'maintenance_status_chart.png'
            plt.savefig(chart_path, dpi=150, bbox_inches='tight', pad_inches=0.05)
            plt.close()
            return chart_path
        except Exception as e:
            self.logger.log(f"❌ Не удалось построить диаграмму: {e}")
            return None

# --- 5.1. Управление обслуженным оборудованием ---
class ServicedEquipmentManager:
    """Класс для управления историей обслуженного оборудования."""
    
    def __init__(self, config: Config, logger: DualLogger):
        self.config = config
        self.logger = logger
        self.serviced_history_file = config.SERVICED_HISTORY_FILE
        self.excel_snapshot_file = config.EXCEL_SNAPSHOT_FILE
    
    # === Работа со снимком дат ТО ===
    def load_snapshot(self) -> Dict[str, str]:
        """Загружает снимок дат ТО из JSON файла.
        
        Returns:
            Словарь: ключ = "sheet:row_number", значение = дата ТО (str)
        """
        try:
            if self.excel_snapshot_file.exists():
                with open(self.excel_snapshot_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    return data.get('snapshot', {})
            return {}
        except Exception as e:
            self.logger.log(f"⚠️ Ошибка при загрузке снимка: {e}")
            return {}
    
    def save_snapshot(self, snapshot: Dict[str, str]) -> None:
        """Сохраняет снимок дат ТО в JSON файл."""
        try:
            self.config.DATA_DIR.mkdir(parents=True, exist_ok=True)
            with open(self.excel_snapshot_file, 'w', encoding='utf-8') as f:
                json.dump({
                    'snapshot': snapshot,
                    'last_update': datetime.now().isoformat()
                }, f, ensure_ascii=False, indent=2)
            self.logger.log(f"✅ Снимок дат ТО сохранён: {len(snapshot)} записей")
        except Exception as e:
            self.logger.log(f"❌ Ошибка при сохранении снимка: {e}")
    
    def create_snapshot(self, excel_file_path: Path) -> Dict[str, str]:
        """Создаёт снимок дат ТО из всех данных Excel.

        Args:
            excel_file_path: Путь к Excel файлу
            
        Returns:
            Словарь: ключ = "sheet:row_number", значение = дата последнего ТО (str)
        """
        snapshot = {}
        
        try:
            # Читаем все листы напрямую из Excel
            for sheet_name in self.config.SHEETS_CONFIG.keys():
                # Читаем данные из Excel
                df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=3, nrows=500)
                
                if len(df.columns) > len(self.config.COLUMN_NAMES):
                    df = df.iloc[:, :len(self.config.COLUMN_NAMES)]
                df.columns = self.config.COLUMN_NAMES
                df = df.dropna(how='all')
                
                # Проходим по всем строкам
                for _, row in df.iterrows():
                    row_num = row.get('№')
                    if pd.isna(row_num):
                        continue
                    
                    try:
                        row_num = int(float(row_num))
                    except (ValueError, TypeError):
                        continue
                    
                    key = f"{sheet_name}:{row_num}"
                    last_date = row.get('Дата последнего ТО')
                    
                    if pd.notna(last_date):
                        if hasattr(last_date, 'strftime'):
                            snapshot[key] = last_date.strftime('%Y-%m-%d')
                        else:
                            snapshot[key] = str(last_date)
                            
        except Exception as e:
            self.logger.log(f"⚠️ Ошибка при создании снимка: {e}")

        return snapshot
    
    # === Работа с историей обслуженного оборудования ===
    def load_serviced_history(self) -> Dict[str, Any]:
        """Загружает историю обслуженного оборудования."""
        try:
            if self.serviced_history_file.exists():
                with open(self.serviced_history_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return {'serviced_equipment': [], 'last_update': None}
        except Exception as e:
            self.logger.log(f"⚠️ Ошибка при загрузке истории: {e}")
            return {'serviced_equipment': [], 'last_update': None}
    
    def save_serviced_history(self, history: Dict[str, Any]) -> None:
        """Сохраняет историю обслуженного оборудования."""
        try:
            self.config.DATA_DIR.mkdir(parents=True, exist_ok=True)
            history['last_update'] = datetime.now().isoformat()
            with open(self.serviced_history_file, 'w', encoding='utf-8') as f:
                json.dump(history, f, ensure_ascii=False, indent=2)
            self.logger.log(f"✅ История обслуженного оборудования сохранена")
        except Exception as e:
            self.logger.log(f"❌ Ошибка при сохранении истории: {e}")
    
    def add_serviced_records(self, new_records: List[Dict[str, Any]]) -> int:
        """Добавляет новые записи в историю обслуженного оборудования.

        Args:
            new_records: Список записей для добавления

        Returns:
            Количество добавленных записей
        """
        if not new_records:
            return 0

        history = self.load_serviced_history()

        # Добавляем новые записи
        for record in new_records:
            history['serviced_equipment'].append(record)
            self.logger.log(f"📝 Добавлено в историю: {record['designation']} ({record['date']})")

        # Очищаем старые записи (храним только за последние SERVICED_HISTORY_DAYS дней)
        cutoff_date = (datetime.now().date() - timedelta(days=self.config.CHART_DAYS)).isoformat()
        original_count = len(history['serviced_equipment'])
        history['serviced_equipment'] = [
            rec for rec in history['serviced_equipment']
            if rec.get('date', '') >= cutoff_date
        ]
        if len(history['serviced_equipment']) < original_count:
            self.logger.log(f"🗑️ Удалено {original_count - len(history['serviced_equipment'])} устаревших записей")

        self.save_serviced_history(history)
        return len(new_records)
    
    def detect_serviced_equipment(self, current_snapshot: Dict[str, str], 
                                  previous_snapshot: Dict[str, str],
                                  all_equipment_data: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Обнаруживает обслуженное оборудование путём сравнения снимков.
        
        Args:
            current_snapshot: Текущий снимок дат ТО
            previous_snapshot: Предыдущий снимок дат ТО
            all_equipment_data: Данные обо всём оборудовании для заполнения записей
                               ключ = "sheet:row_number"
                               
        Returns:
            Список новых записей об обслуживании
        """
        new_records = []
        today = datetime.now().date()
        today_str = today.isoformat()
        timestamp = datetime.now().isoformat()
        
        # Ищем записи, где дата ТО изменилась
        for key, current_date in current_snapshot.items():
            previous_date = previous_snapshot.get(key)
            
            # Если предыдущей даты нет (первый запуск) или даты разные и текущая больше
            if previous_date is None:
                continue  # Пропускаем записи, которых не было в предыдущем снимке
            
            if previous_date != current_date and current_date > previous_date:
                # Получаем данные об оборудовании
                equipment = all_equipment_data.get(key, {})
                
                if equipment:
                    record = {
                        'date': today_str,
                        'timestamp': timestamp,
                        'sheet': equipment.get('sheet', ''),
                        'row': equipment.get('row', ''),
                        'designation': equipment.get('designation', ''),
                        'name': equipment.get('name', ''),
                        'object': equipment.get('object', '')
                    }
                    new_records.append(record)
        
        return new_records
    
    def get_serviced_last_days(self, days: int = 7) -> List[Dict[str, Any]]:
        """Получает записи об обслуживании за последние N дней.
        
        Args:
            days: Количество дней
            
        Returns:
            Список записей за указанный период
        """
        history = self.load_serviced_history()
        cutoff_date = (datetime.now().date() - timedelta(days=days)).isoformat()
        
        filtered = [
            record for record in history['serviced_equipment']
            if record.get('date', '') >= cutoff_date
        ]
        
        # Сортируем по дате (свежие сверху)
        return sorted(filtered, key=lambda x: x.get('date', ''), reverse=True)
    
    def get_serviced_equipment_data(self, excel_file_path: Path) -> Dict[str, Dict[str, Any]]:
        """Извлекает данные обо всём оборудовании для использования при детектировании.

        Args:
            excel_file_path: Путь к Excel файлу
            
        Returns:
            Словарь: ключ = "sheet:row_number", значение = данные оборудования
        """
        equipment_data = {}
        
        try:
            # Читаем все листы напрямую из Excel
            for sheet_name in self.config.SHEETS_CONFIG.keys():
                df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=3, nrows=500)
                
                if len(df.columns) > len(self.config.COLUMN_NAMES):
                    df = df.iloc[:, :len(self.config.COLUMN_NAMES)]
                df.columns = self.config.COLUMN_NAMES
                df = df.dropna(how='all')
                
                for _, row in df.iterrows():
                    row_num = row.get('№')
                    if pd.isna(row_num):
                        continue
                    
                    try:
                        row_num = int(float(row_num))
                    except (ValueError, TypeError):
                        continue
                    
                    key = f"{sheet_name}:{row_num}"
                    
                    equipment_data[key] = {
                        'sheet': sheet_name,
                        'row': row_num,
                        'designation': row.get('Обозначение', ''),
                        'name': row.get('Наименование', ''),
                        'object': row.get('Объект', '')
                    }
                    
        except Exception as e:
            self.logger.log(f"⚠️ Ошибка при извлечении данных оборудования: {e}")
        
        return equipment_data

    def record_manual_service(self, excel_file_path: Path,
                              items: List[Tuple[str, int]]) -> int:
        """Записывает в журнал факты ручной отметки «обслужено» через web-интерфейс.

        Для каждого (sheet_name, row_number) формирует запись, добавляет её в
        serviced_history.json (с дедупликацией по sheet+row+date за сегодняшний день)
        и обновляет excel_snapshot.json, чтобы плановый запуск maintenance_alert.py
        не породил дубль той же операции.

        Args:
            excel_file_path: Путь к Excel-файлу (для извлечения метаданных оборудования).
            items: Список кортежей (sheet_name, row_number).

        Returns:
            Количество фактически добавленных (недублирующих) записей.
        """
        if not items:
            return 0

        equipment_data = self.get_serviced_equipment_data(excel_file_path)
        history = self.load_serviced_history()
        today_str = datetime.now().date().isoformat()
        timestamp = datetime.now().isoformat()

        # Индекс существующих записей за сегодня для дедупликации
        existing_today = {
            (rec.get('sheet', ''), str(rec.get('row', '')))
            for rec in history.get('serviced_equipment', [])
            if rec.get('date', '') == today_str
        }

        added = 0
        for sheet_name, row_number in items:
            try:
                row_num_int = int(row_number)
            except (ValueError, TypeError):
                continue

            dedup_key = (sheet_name, str(row_num_int))
            if dedup_key in existing_today:
                self.logger.log(
                    f"ℹ️ Пропущено (уже в журнале за сегодня): {sheet_name} / №{row_num_int}"
                )
                continue

            key = f"{sheet_name}:{row_num_int}"
            equipment = equipment_data.get(key, {})

            record = {
                'date': today_str,
                'timestamp': timestamp,
                'sheet': equipment.get('sheet', sheet_name),
                'row': equipment.get('row', row_num_int),
                'designation': equipment.get('designation', ''),
                'name': equipment.get('name', ''),
                'object': equipment.get('object', '')
            }
            history.setdefault('serviced_equipment', []).append(record)
            existing_today.add(dedup_key)
            added += 1
            self.logger.log(
                f"📝 Ручная отметка в журнал: {record['designation']} ({record['object']})"
            )

        # Очищаем записи старше CHART_DAYS (консистентно с add_serviced_records)
        cutoff_date = (datetime.now().date() - timedelta(days=self.config.CHART_DAYS)).isoformat()
        original_count = len(history['serviced_equipment'])
        history['serviced_equipment'] = [
            rec for rec in history['serviced_equipment']
            if rec.get('date', '') >= cutoff_date
        ]
        if len(history['serviced_equipment']) < original_count:
            self.logger.log(
                f"🗑️ Удалено {original_count - len(history['serviced_equipment'])} устаревших записей"
            )

        if added > 0:
            self.save_serviced_history(history)

        # Обновляем снимок дат ТО, чтобы scheduled-запуск не создал дубль
        try:
            current_snapshot = self.create_snapshot(excel_file_path)
            self.save_snapshot(current_snapshot)
        except Exception as e:
            self.logger.log(f"⚠️ Не удалось обновить снимок после ручной отметки: {e}")

        return added

# --- 6. Генерация отчета ---
class ReportGenerator:
    """Класс для генерации HTML-отчета."""
    def __init__(self, config: Config, logger: DualLogger, maintenance_checker: MaintenanceChecker, 
                 statistics_manager: StatisticsManager, serviced_equipment_manager: ServicedEquipmentManager = None):
        self.config = config
        self.logger = logger
        self.maintenance_checker = maintenance_checker
        self.statistics_manager = statistics_manager
        self.serviced_equipment_manager = serviced_equipment_manager

    def _create_serviced_email_block(self, serviced_records: List[Dict[str, Any]]) -> str:
        """Создаёт HTML-блок для обслуженного оборудования за последние 7 дней."""
        if not serviced_records:
            return ""
        
        # Группируем по датам
        by_date = {}
        for record in serviced_records:
            date_str = record.get('date', '')
            if date_str not in by_date:
                by_date[date_str] = []
            by_date[date_str].append(record)
        
        html_parts = []
        html_parts.append("<br/>")
        html_parts.append("<div style='background-color: #f8f9fa; border-radius: 8px; padding: 15px; border-left: 4px solid #18bc9c; margin-bottom: 20px;'>")
        html_parts.append("<div style='font-size: 15px; font-weight: bold; color: #2c3e50; margin-bottom: 10px;'>📝 Обслужено за последние 7 дней</div>")
        
        for date_str in sorted(by_date.keys(), reverse=True):
            records = by_date[date_str]
            # Форматируем дату
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                formatted_date = date_obj.strftime('%d.%m.%Y')
            except:
                formatted_date = date_str
            
            html_parts.append(f"<div style='font-size: 13px; font-weight: bold; color: #18bc9c; margin-top: 10px; margin-bottom: 5px;'>📅 {formatted_date} ({len(records)} ед.)</div>")
            html_parts.append("<table style='width: 100%; border-collapse: collapse; font-size: 12px; table-layout: fixed;'>")
            html_parts.append("<thead><tr style='background-color: #e8f5e9;'>")
            html_parts.append("<th style='padding: 6px; border: 1px solid #cfd8dc; text-align: left; width: 150px;'>Объект</th>")
            html_parts.append("<th style='padding: 6px; border: 1px solid #cfd8dc; text-align: left; width: 200px;'>Обозначение</th>")
            html_parts.append("<th style='padding: 6px; border: 1px solid #cfd8dc; text-align: left;'>Наименование</th>")
            html_parts.append("</tr></thead>")
            html_parts.append("<tbody>")

            for rec in records:
                html_parts.append("<tr>")
                html_parts.append(f"<td style='padding: 6px; border: 1px solid #cfd8dc; color: #2c3e50; width: 150px;'>{rec.get('object', '')}</td>")
                html_parts.append(f"<td style='padding: 6px; border: 1px solid #cfd8dc; font-weight: bold; color: #2c3e50; width: 200px;'>{rec.get('designation', '')}</td>")
                html_parts.append(f"<td style='padding: 6px; border: 1px solid #cfd8dc; color: #2c3e50;'>{rec.get('name', '')}</td>")
                html_parts.append("</tr>")
            
            html_parts.append("</tbody></table>")
        
        html_parts.append("</div>")
        return "".join(html_parts)

    def create_body(self, urgent_items: List[pd.DataFrame],
                    warning_items: List[pd.DataFrame],
                    total_records: int,
                    status_counts: Dict[str, int],
                    recalc_success: bool = True) -> Tuple[str, Optional[Path]]:
        """Создает HTML-тело письма."""
        # Вычисляем процент необслуженного оборудования
        unserviced_count = status_counts[self.config.STATUS_URGENT] #+ status_counts[self.config.STATUS_WARNING]
        unserviced_percentage = (unserviced_count / total_records * 100) if total_records > 0 else 0
        html_parts: List[str] = []
        
        # Обертка для всей почты, чтобы ограничить ширину и выровнять элементы
        html_parts.append("<div style='width: 100%; max-width: 1200px; font-family: Segoe UI, Tahoma, Geneva, Verdana, sans-serif;'>")
        
        # Предупреждение о неудачном пересчете формул
        if not recalc_success:
            html_parts.append(
                f"""
                <div style="background-color: #ff6b6b; border-radius: 8px; padding: 15px; border-left: 5px solid #e74c3c;
                            color: white; margin-bottom: 20px; display: flex; align-items: center;">
                    <div style="margin-right: 15px;">
                        <img src="cid:app_icon_alert" alt="Иконка приложения" style="width: 86px; height: 86px; border-radius: 8px;">
                    </div>
                    <div style="text-align: left;">
                        <div style="font-size: 16px; font-weight: bold; margin-bottom: 10px;">⚠️ ВНИМАНИЕ! ТАБЛИЦА ОТКРЫТА! ⚠️</div>
                        <div style="font-size: 16px; line-height: 1.4;">
                            Перерасчёт графика обслуживания невозможен!<br/>
                            Закройте таблицу чтобы восстановить расчёты, или живите дальше в проклятом мире, который сами и создали!
                        </div>
                    </div>
                </div>
                """
            )
        # Верхняя сводка - компактный вариант с названиями над цифрами #2c3e50 #2c3e50
        html_parts.append(
            f"""
            <div style="background-color: #2c3e50; border-radius: 8px; padding: 15px; border-left: 4px solid #18bc9c;
                        color: white;">
                <div style="display: flex; justify-content: space-around; text-align: center; flex-wrap: wrap;">
                    <div style="margin: 5px; ">
                        <div style="font-size: 12px; color: #ffd6d6; margin-bottom: 3px;">🚨 ОБСЛУЖИТЬ</div>
                        <div style="font-size: 20px; font-weight: bold; color: #ff6b6b;">{status_counts[self.config.STATUS_URGENT]} ({unserviced_percentage:.1f}%) </div>
                    </div>
                    <div style="margin: 5px; margin-left: 20px;">
                        <div style="font-size: 12px; color: #ffe082; margin-bottom: 3px;">⚠️ Внимание</div>
                        <div style="font-size: 20px; font-weight: bold; color: #ffd54f;">{status_counts[self.config.STATUS_WARNING]}</div>
                    </div>
                    <div style="margin: 5px; margin-left: 20px;">
                        <div style="font-size: 12px; color: #18bc9c; margin-bottom: 3px;">✅ Не требуется</div>
                        <div style="font-size: 20px; font-weight: bold; color: #18bc9c;">{status_counts[self.config.STATUS_OK]}</div>
                    </div>
                    <div style="margin: 5px; margin-left: 20px;">
                        <div style="font-size: 12px; color: #bbdefb; margin-bottom: 3px;">📊 Всего</div>
                        <div style="font-size: 20px; font-weight: bold; color: #4fc3f7;">{total_records}</div>
                    </div>
                    <div style="margin-left: 25px;">
                        <a href="http://10.100.59.40:5940/" title="Перейти в панель управления">
                            <img src="cid:app_icon" alt="Иконка приложения" style="width: 52px; height: 52px; border-radius: 8px; border: none;">
                        </a>
                    </div>
                </div>
            </div>
            <br/>
            """
        )
        # Создаем диаграмму
        chart_path = self.statistics_manager.create_chart()
        # Вставляем диаграмму ПЕРЕД секцией срочных работ
        if chart_path and Path(chart_path).exists():
            html_parts.append(
                (
                    "<div style='margin-bottom: 20px;'>"
                    "<img src=\"cid:status_chart\" alt=\"Диаграмма\" style='width: 100%; display: block; border-radius: 8px;'/>"
                    "</div>"
                )
            )
        # Срочные и внимание элементы в раздельных таблицах
        if urgent_items or warning_items:
            total_urgent = sum(len(df) for df in urgent_items) if urgent_items else 0
            total_warning = sum(len(df) for df in warning_items) if warning_items else 0
            
            # 1. Срочные элементы
            if urgent_items:
                html_parts.append(f"<div><strong style='color:#e74c3c;'>🚨 ОБСЛУЖИТЬ (записей: {total_urgent}):</strong></div>")
                html_parts.append("<hr style='background-color: #e74c3c; height: 2px; border: none;' />")
                html_parts.append("""
                <table style='width:100%; border-collapse:collapse; font-size:13px; margin-top:10px; margin-bottom:20px;'>
                    <thead>
                        <tr style='background-color:#2c3e50; color:white;'>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Объект</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Наименование</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Обозначение</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Место расположения</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Работы</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Инт. ТО</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Дата ТО</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Дата след. ТО</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Статус</th>
                        </tr>
                    </thead>
                    <tbody>
                """)
                
                combined_urgent = pd.concat(urgent_items).sort_values(by='Объект')
                color_index = 0
                for _, item in combined_urgent.iterrows():
                    bg_color = '#ffffff' if color_index % 2 == 0 else '#f9f9f9'
                    html_parts.append(self.maintenance_checker.format_item_table_row(item, bg_color))
                    color_index += 1
                
                html_parts.append("</tbody></table>")
            
            # 2. Элементы требующие внимания
            if warning_items:
                html_parts.append(f"<div><strong style='color:#f39c12;'>⚠️ ВНИМАНИЕ! Приближается срок обслуживания (записей: {total_warning}):</strong></div>")
                html_parts.append("<hr style='background-color: #f39c12; height: 2px; border: none;' />")
                html_parts.append("""
                <table style='width:100%; border-collapse:collapse; font-size:13px; margin-top:10px; margin-bottom:20px;'>
                    <thead>
                        <tr style='background-color:#2c3e50; color:white;'>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Объект</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Наименование</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Обозначение</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Место расположения</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Работы</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Инт. ТО</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Дата ТО</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Дата след. ТО</th>
                            <th style='padding:10px; text-align:left; border:1px solid #cfd8dc;'>Статус</th>
                        </tr>
                    </thead>
                    <tbody>
                """)
                
                combined_warning = pd.concat(warning_items).sort_values(by='Объект')
                color_index = 0
                for _, item in combined_warning.iterrows():
                    bg_color = '#fffdf0' if color_index % 2 == 0 else '#fff9e6'
                    html_parts.append(self.maintenance_checker.format_item_table_row(item, bg_color))
                    color_index += 1
                
                html_parts.append("</tbody></table>")

            html_parts.append("<br/>")
        
        # Блок обслуженного оборудования за последние 7 дней
        if self.serviced_equipment_manager:
            serviced_records = self.serviced_equipment_manager.get_serviced_last_days(7)
            if serviced_records:
                html_parts.append(self._create_serviced_email_block(serviced_records))
        
        # нижняя часть письма
        html_parts.append(
            f"""
            <br/>
            <div style="background-color: #EFF2F6; border-left: 4px solid #18bc9c; 
                        padding: 12px; margin-top: 20px; font-size: 11px; color: #333;">
                <div style="margin-bottom: 8px;">
                    <span style="font-weight: bold;color:#2c3e50;">🔧 Скрипт рассылки уведомлений об обслуживании оборудования АСУТП</span> 
                    <span style="float: right; background-color: #18bc9c; color: white; 
                                padding: 2px 8px; border-radius: 10px; font-size: 10px;">
                        v{self.config.VERSION} от {self.config.RELEASE_DATE}<br/> semonoff@gmail.com
                    </span>
                    <span style="float: right; margin-right: 8px "> 
                        <img src="cid:app_icon" alt="Иконка приложения" style="width: 32px; height: 32px; border-radius: 8px;">
                    </span>
                </div>
                <div style="line-height: 1.4;">
                    <span style="color: #2c3e50;">📂 Файлы на сервере ASUTP-FILES-SRV01:</span><br/>
                    <span style="margin-left: 15px;">📊 Таблица:</span> <code>{self.config.get_excel_file_path()}</code><br/>
                    <span style="margin-left: 15px;">🐍 Скрипт:</span> <code>{Path(__file__).resolve()}</code> <br/>
                    <span style="">⏰ Запуск:</span> Ежедневно из Task Scheduler, правило: <code>maintenance_alert.py</code><br/>
                    <span style="">🖥️ Панель управления:</span> <a href="http://10.100.59.40:5940/" style="color: #18bc9c; text-decoration: none;">http://10.100.59.40:5940/</a><br/>
                    <span style="">🌐 Исходный код:</span> <a href="https://github.com/SemonoffArt/maintenance_alert" style="color: #18bc9c; text-decoration: none;">GitHub репозиторий</a><br/>
                    <span style="">📧 Получатели ({len(self.config.RECIPIENTS)}):</span> {', '.join(self.config.RECIPIENTS)}<br/>
                    <div style="text-align: right; margin-top: 5px; color: #2c3e50; font-size: 10px;">
                        Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}
                    </div>
                </div>
            </div>
            """
        )
        html_parts.append("</div>") # Закрытие общей обертки
        html_body = "".join(html_parts)
        return html_body, chart_path

# --- 7. Отправка почты ---
class EmailSender:
    """Класс для отправки email."""
    def __init__(self, config: Config, logger: DualLogger):
        self.config = config
        self.logger = logger

    def send(self, html_body: str, recipients: List[str], chart_path: Optional[Path] = None, attachment_path: Optional[Path] = None) -> bool:
        """Отправляет email через SMTP."""
        try:
            # Создаем сообщение
            msg = MIMEMultipart('related')
            msg['From'] = self.config.SENDER_EMAIL
            msg['To'] = ", ".join(recipients)
            msg['Subject'] = "🔔 Напоминание о техническом обслуживании оборудования"
            alternative = MIMEMultipart('alternative')
            msg.attach(alternative)
            # Добавляем HTML-контент и изображение при наличии
            if chart_path and Path(chart_path).exists():
                alternative.attach(MIMEText(html_body, 'html', 'utf-8'))
                with open(chart_path, 'rb') as img_file:
                    img = MIMEImage(img_file.read())
                    img.add_header('Content-ID', '<status_chart>')
                    img.add_header('Content-Disposition', 'inline', filename=Path(chart_path).name)
                    msg.attach(img)
            else:
                alternative.attach(MIMEText(html_body, 'html', 'utf-8'))
            # Добавляем иконку приложения
            icon_path = self.config.DATA_DIR / "manky.png"
            if icon_path.exists():
                with open(icon_path, 'rb') as icon_file:
                    icon = MIMEImage(icon_file.read())
                    icon.add_header('Content-ID', '<app_icon>')
                    icon.add_header('Content-Disposition', 'inline', filename='manky.png')
                    msg.attach(icon)
            # Добавляем иконку приложения
            icon_path = self.config.DATA_DIR / "manky_alert.png"
            if icon_path.exists():
                with open(icon_path, 'rb') as icon_file:
                    icon = MIMEImage(icon_file.read())
                    icon.add_header('Content-ID', '<app_icon_alert>')
                    icon.add_header('Content-Disposition', 'inline', filename='manky_alert.png')
                    msg.attach(icon)
            
            # Добавляем вложение с файлом maintenance_data.xlsx
            if attachment_path and attachment_path.exists():
                with open(attachment_path, 'rb') as attachment_file:
                    attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    attachment.set_payload(attachment_file.read())
                    encoders.encode_base64(attachment)
                    attachment.add_header(
                        'Content-Disposition',
                        f'attachment; filename="{attachment_path.name}"'
                    )
                    msg.attach(attachment)
                    self.logger.log(f"📎 Прикреплен файл: {attachment_path.name}")
            # Подключаемся к SMTP серверу и отправляем письмо
            server = smtplib.SMTP(self.config.SMTP_SERVER, self.config.SMTP_PORT)
            # Не используем starttls() для порта 25 без шифрования
            server.sendmail(self.config.SENDER_EMAIL, recipients, msg.as_string())
            server.quit()
            self.logger.log(f"✅ Письмо успешно отправлено {len(recipients)} получателям")
            return True
        except Exception as e:
            self.logger.log(f"❌ Ошибка при отправке письма: {e}")
            return False

# --- 8. Основная логика ---
class MaintenanceAlertApp:
    """Основной класс приложения."""
    def __init__(self):
        self.config = Config()
        self.logger = DualLogger(self.config.LOG_FILE)
        self.excel_handler = ExcelHandler(self.config, self.logger)
        self.maintenance_checker = MaintenanceChecker(self.config, self.logger)
        self.statistics_manager = StatisticsManager(self.config, self.logger)
        self.serviced_equipment_manager = ServicedEquipmentManager(self.config, self.logger)
        self.report_generator = ReportGenerator(self.config, self.logger, self.maintenance_checker, 
                                                 self.statistics_manager, self.serviced_equipment_manager)
        self.email_sender = EmailSender(self.config, self.logger)

    def show_version(self):
        """Отображает информацию о версии программы."""
        self.logger.log(f"🔧 Система уведомлений о техническом обслуживании v{self.config.VERSION}")
        self.logger.log(f"📅 Дата выпуска: {self.config.RELEASE_DATE}")
        self.logger.log(f"🐍 Python: {sys.version.split()[0]}")
        self.logger.log_separator()

    def run(self):
        """Основная функция программы."""
        self.show_version()
        self.logger.log("🚀 ПРОГРАММА ЗАПУЩЕНА")
        self.logger.log("Начинаем проверку графика технического обслуживания...")
        self.logger.log(f"Получатели: {', '.join(self.config.RECIPIENTS)}")

        alarm_items, warning_items, total_records, status_counts, recalc_success = self.excel_handler.read_data()

        self.logger.log("\n" + "="*60)
        self.logger.log("📈 ОБНОВЛЕНИЕ СТАТИСТИКИ ОБСЛУЖИВАНИЯ")
        self.logger.log_separator()
        self.statistics_manager.update_statistics(alarm_items, warning_items, total_records, status_counts)
        self.logger.log_separator()

        # === DETECTION OF SERVICED EQUIPMENT ===
        self.logger.log("\n" + "="*60)
        self.logger.log("🔍 ПОИСК ОБСЛУЖЕННОГО ОБОРУДОВАНИЯ")
        self.logger.log_separator()

        # Загружаем предыдущий снимок
        previous_snapshot = self.serviced_equipment_manager.load_snapshot()
        self.logger.log(f"📊 Загружен предыдущий снимок: {len(previous_snapshot)} записей")

        # Получаем путь к Excel файлу (используется после read_data())
        excel_file = self.excel_handler.get_last_excel_file_path()

        # Создаём текущий снимок и получаем данные оборудования
        current_snapshot = self.serviced_equipment_manager.create_snapshot(excel_file)
        equipment_data = self.serviced_equipment_manager.get_serviced_equipment_data(excel_file)
        self.logger.log(f"📊 Создан текущий снимок: {len(current_snapshot)} записей")
        self.logger.log(f"📊 Получены данные оборудования: {len(equipment_data)} записей")

        # Обнаруживаем изменения
        new_records = self.serviced_equipment_manager.detect_serviced_equipment(
            current_snapshot, previous_snapshot, equipment_data
        )

        if new_records:
            self.logger.log(f"✅ Обнаружено {len(new_records)} ед. обслуженного оборудования:")
            for rec in new_records:
                self.logger.log(f"   - {rec['designation']} ({rec['object']})")
            # Добавляем в историю
            self.serviced_equipment_manager.add_serviced_records(new_records)
        else:
            self.logger.log("ℹ️ Новых записей об обслуживании не обнаружено")

        # Сохраняем текущий снимок для следующего сравнения
        self.serviced_equipment_manager.save_snapshot(current_snapshot)
        self.logger.log_separator()
        # === END OF SERVICED DETECTION ===

        total_alarm = sum(len(df) for df in alarm_items) if alarm_items else 0
        total_warning = sum(len(df) for df in warning_items) if warning_items else 0
        self.logger.log(f"\nИтого найдено:")
        self.logger.log(f"  {self.config.STATUS_URGENT}: {total_alarm}")
        self.logger.log(f"  {self.config.STATUS_WARNING}: {total_warning}")

        if total_alarm == 0 and total_warning == 0:
            self.logger.log("Нет срочных напоминаний. Все оборудование в порядке.")
            return

        email_body, chart_path = self.report_generator.create_body(
            alarm_items, warning_items, total_records, status_counts, recalc_success
        )
        self.logger.log("\nСформировано письмо:")

        # Генерируем файл maintenance_data.xlsx с данными для обслуживания
        maintenance_data_file = None
        if alarm_items:  # Создаем файл только если есть срочные задачи
            self.logger.log("📝 Генерируем файл maintenance_data.xlsx...")
            maintenance_data_file = self.excel_handler.generate_maintenance_data_file(alarm_items)
            if maintenance_data_file:
                self.logger.log(f"✅ Файл с данными для обслуживания готов: {maintenance_data_file.name}")
            else:
                self.logger.log("⚠️ Не удалось создать файл maintenance_data.xlsx")

        self.logger.log(f"\nОтправляем письмо {len(self.config.RECIPIENTS)} получателям...")
        if self.email_sender.send(email_body, self.config.RECIPIENTS, chart_path, maintenance_data_file):
            self.logger.log("Письма отправлены успешно")

            # Удаляем временный файл maintenance_data.xlsx после успешной отправки
            if maintenance_data_file and maintenance_data_file.exists():
                try:
                    maintenance_data_file.unlink()
                    self.logger.log(f"🗑️ Временный файл удален: {maintenance_data_file.name}")
                except Exception as e:
                    self.logger.log(f"⚠️ Не удалось удалить временный файл {maintenance_data_file.name}: {e}")
        else:
            self.logger.log("Не удалось отправить письма")
        self.logger.log("\n\n\n")

def main():
    """Точка входа в программу."""
    app = MaintenanceAlertApp()
    app.run()

if __name__ == "__main__":
    main()